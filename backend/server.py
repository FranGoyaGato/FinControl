from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, File, Query, Depends, Header
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import hashlib
import io
import csv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, numbers
import xlrd
from decimal import Decimal
import re
import bcrypt
import jwt
from pymongo.errors import DuplicateKeyError


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# ============ AUTH UTILITIES ============

JWT_ALGORITHM = "HS256"
JWT_EXPIRE_DAYS = 30


def _jwt_secret() -> str:
    return os.environ['JWT_SECRET']


def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')


def verify_password(plain: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(plain.encode('utf-8'), hashed.encode('utf-8'))
    except (ValueError, TypeError):
        return False


def create_access_token(user_id: str, email: str) -> str:
    payload = {
        'sub': user_id,
        'email': email,
        'exp': datetime.now(timezone.utc) + timedelta(days=JWT_EXPIRE_DAYS),
        'iat': datetime.now(timezone.utc),
        'type': 'access',
    }
    return jwt.encode(payload, _jwt_secret(), algorithm=JWT_ALGORITHM)


async def get_current_user(authorization: Optional[str] = Header(default=None)) -> dict:
    if not authorization or not authorization.lower().startswith('bearer '):
        raise HTTPException(status_code=401, detail='No autenticado')
    token = authorization.split(' ', 1)[1].strip()
    try:
        payload = jwt.decode(token, _jwt_secret(), algorithms=[JWT_ALGORITHM])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail='Sesión expirada')
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail='Token inválido')
    if payload.get('type') != 'access':
        raise HTTPException(status_code=401, detail='Token inválido')
    user = await db.users.find_one({'id': payload['sub']}, {'_id': 0, 'password_hash': 0})
    if not user:
        raise HTTPException(status_code=401, detail='Usuario no encontrado')
    return user


# ============ AUTH ROUTER (public) ============

auth_router = APIRouter(prefix="/api/auth")


class LoginRequest(BaseModel):
    email: EmailStr
    password: str


class ChangePasswordRequest(BaseModel):
    current_password: str
    new_password: str = Field(min_length=8)


@auth_router.post("/login")
async def login(payload: LoginRequest):
    email = payload.email.lower().strip()
    user = await db.users.find_one({'email': email})
    if not user or not verify_password(payload.password, user.get('password_hash', '')):
        raise HTTPException(status_code=401, detail='Email o contraseña incorrectos')
    token = create_access_token(user['id'], user['email'])
    return {
        'token': token,
        'user': {'id': user['id'], 'email': user['email'], 'name': user.get('name', '')},
    }


@auth_router.get("/me")
async def me(current_user: dict = Depends(get_current_user)):
    return current_user


@auth_router.post("/logout")
async def logout(current_user: dict = Depends(get_current_user)):
    # Stateless JWT — client discards its token. Endpoint kept for API symmetry.
    return {'ok': True}


@auth_router.post("/change-password")
async def change_password(body: ChangePasswordRequest, current_user: dict = Depends(get_current_user)):
    user_doc = await db.users.find_one({'id': current_user['id']})
    if not user_doc or not verify_password(body.current_password, user_doc.get('password_hash', '')):
        raise HTTPException(status_code=400, detail='Contraseña actual incorrecta')
    await db.users.update_one(
        {'id': current_user['id']},
        {'$set': {'password_hash': hash_password(body.new_password),
                  'password_updated_at': datetime.now(timezone.utc).isoformat()}},
    )
    return {'ok': True}


# ============ PROTECTED API ROUTER ============

# Create a router with the /api prefix — all routes require a valid Bearer token.
api_router = APIRouter(prefix="/api", dependencies=[Depends(get_current_user)])


# ============ SEED ADMIN ============

async def seed_admin():
    """Create the single admin user on startup if it doesn't exist. Idempotent
    even under a race between two concurrent startups: the unique index on
    users.email is the source of truth, so a DuplicateKeyError just means the
    user is already there — safe to ignore."""
    admin_email = os.environ['ADMIN_EMAIL'].lower().strip()
    admin_password = os.environ['ADMIN_PASSWORD']
    existing = await db.users.find_one({'email': admin_email})
    if existing is not None:
        return
    try:
        await db.users.insert_one({
            'id': str(uuid.uuid4()),
            'email': admin_email,
            'password_hash': hash_password(admin_password),
            'name': 'Admin',
            'role': 'admin',
            'created_at': datetime.now(timezone.utc).isoformat(),
        })
        logging.info(f"Seeded admin user {admin_email}")
    except DuplicateKeyError:
        logging.info(f"Admin user {admin_email} already exists (race); skipping seed")
    # If it exists we DO NOT reset the hash — the user may have rotated it via /change-password.


@app.on_event("startup")
async def startup_seed_and_indexes():
    try:
        await db.users.create_index('email', unique=True)
    except Exception as e:
        logging.warning(f"Users email index: {e}")
    await seed_admin()



# ============ MODELS ============

class Account(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    type: str = 'bank'  # 'bank'
    currency: str = 'EUR'
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class AccountCreate(BaseModel):
    name: str
    type: str = 'bank'
    currency: str = 'EUR'

class CreditCard(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    last4: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CreditCardCreate(BaseModel):
    name: str
    last4: Optional[str] = None

class Category(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str

class CategoryCreate(BaseModel):
    name: str

class Subcategory(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    category_id: str
    name: str

class SubcategoryCreate(BaseModel):
    category_id: str
    name: str

class Transaction(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    account_id: str
    date: str  # ISO format YYYY-MM-DD
    concept: str
    amount: float
    type: str  # 'income' or 'expense'
    category_id: Optional[str] = None
    subcategory_id: Optional[str] = None
    dedup_hash: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class TransactionCreate(BaseModel):
    account_id: str
    date: str
    concept: str
    amount: float
    type: str
    category_id: Optional[str] = None
    subcategory_id: Optional[str] = None

class CardTransaction(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    card_id: str
    date: str  # ISO format YYYY-MM-DD
    concept: str
    amount: float
    category_id: Optional[str] = None
    subcategory_id: Optional[str] = None
    dedup_hash: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CardTransactionCreate(BaseModel):
    card_id: str
    date: str
    concept: str
    amount: float
    category_id: Optional[str] = None
    subcategory_id: Optional[str] = None

class Rule(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    source: str  # 'bank' or 'card'
    contains: str  # text pattern
    sign: Optional[str] = None  # '+' or '-' or null
    category_id: str
    subcategory_id: Optional[str] = None
    priority: int = 0
    active: bool = True

class RuleCreate(BaseModel):
    source: str
    contains: str
    sign: Optional[str] = None
    category_id: str
    subcategory_id: Optional[str] = None
    priority: int = 0
    active: bool = True

class Settings(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    currency_symbol: str = '€'
    locale: str = 'es-ES'

class SettingsUpdate(BaseModel):
    currency_symbol: str = '€'
    locale: str = 'es-ES'


# ============ HELPER FUNCTIONS ============

def generate_dedup_hash(identifier: str, date: str, concept: str, amount: float) -> str:
    """Generate SHA1 hash for deduplication"""
    raw = f"{identifier}|{date}|{concept}|{amount}"
    return hashlib.sha1(raw.encode()).hexdigest()

def normalize_date(date_str: str) -> str:
    """Convert DD/MM/YYYY to YYYY-MM-DD"""
    if '/' in date_str:
        parts = date_str.strip().split('/')
        if len(parts) == 3:
            day, month, year = parts
            return f"{year.zfill(4)}-{month.zfill(2)}-{day.zfill(2)}"
    return date_str

def detect_separator(content: str) -> str:
    """Detect CSV separator (comma or semicolon)"""
    first_line = content.split('\n')[0]
    if ';' in first_line:
        return ';'
    return ','

async def apply_rules(concept: str, amount: float, source: str) -> tuple:
    """Apply categorization rules. Returns (category_id, subcategory_id)"""
    rules = await db.rules.find({
        'source': source,
        'active': True
    }).sort('priority', -1).to_list(1000)
    
    for rule in rules:
        if rule['contains'].lower() in concept.lower():
            # Check sign if specified
            if rule.get('sign'):
                if rule['sign'] == '+' and amount <= 0:
                    continue
                if rule['sign'] == '-' and amount >= 0:
                    continue
            return (rule['category_id'], rule.get('subcategory_id'))
    
    return (None, None)

def format_currency_es(value: float) -> str:
    """Format number as EUR es-ES"""
    formatted = f"{value:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    return f"{formatted} €"


# ============ ROUTES ============

@api_router.get("/")
async def root():
    return {"message": "Finance Control API"}


# --- ACCOUNTS ---
@api_router.post("/accounts", response_model=Account)
async def create_account(input: AccountCreate):
    account = Account(**input.model_dump())
    doc = account.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.accounts.insert_one(doc)
    return account

@api_router.get("/accounts", response_model=List[Account])
async def get_accounts():
    accounts = await db.accounts.find({}, {"_id": 0}).to_list(1000)
    for acc in accounts:
        if isinstance(acc.get('created_at'), str):
            acc['created_at'] = datetime.fromisoformat(acc['created_at'])
    return accounts

@api_router.delete("/accounts/{account_id}")
async def delete_account(account_id: str):
    result = await db.accounts.delete_one({"id": account_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Account not found")
    # Also delete associated transactions
    await db.transactions.delete_many({"account_id": account_id})
    return {"message": "Account deleted"}


# --- CREDIT CARDS ---
@api_router.post("/credit-cards", response_model=CreditCard)
async def create_credit_card(input: CreditCardCreate):
    card = CreditCard(**input.model_dump())
    doc = card.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.credit_cards.insert_one(doc)
    return card

@api_router.get("/credit-cards", response_model=List[CreditCard])
async def get_credit_cards():
    cards = await db.credit_cards.find({}, {"_id": 0}).to_list(1000)
    for card in cards:
        if isinstance(card.get('created_at'), str):
            card['created_at'] = datetime.fromisoformat(card['created_at'])
    return cards

@api_router.delete("/credit-cards/{card_id}")
async def delete_credit_card(card_id: str):
    result = await db.credit_cards.delete_one({"id": card_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Card not found")
    # Also delete associated transactions
    await db.card_transactions.delete_many({"card_id": card_id})
    return {"message": "Card deleted"}


# --- CATEGORIES ---
@api_router.post("/categories", response_model=Category)
async def create_category(input: CategoryCreate):
    category = Category(**input.model_dump())
    await db.categories.insert_one(category.model_dump())
    return category

@api_router.get("/categories", response_model=List[Category])
async def get_categories():
    categories = await db.categories.find({}, {"_id": 0}).to_list(1000)
    return categories

@api_router.put("/categories/{category_id}", response_model=Category)
async def update_category(category_id: str, input: CategoryCreate):
    result = await db.categories.find_one({"id": category_id}, {"_id": 0})
    if not result:
        raise HTTPException(status_code=404, detail="Category not found")
    
    await db.categories.update_one({"id": category_id}, {"$set": {"name": input.name}})
    result['name'] = input.name
    return Category(**result)

@api_router.delete("/categories/{category_id}")
async def delete_category(category_id: str):
    result = await db.categories.delete_one({"id": category_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Category not found")
    # Also delete subcategories
    await db.subcategories.delete_many({"category_id": category_id})
    return {"message": "Category deleted"}


# --- SUBCATEGORIES ---
@api_router.post("/subcategories", response_model=Subcategory)
async def create_subcategory(input: SubcategoryCreate):
    subcategory = Subcategory(**input.model_dump())
    await db.subcategories.insert_one(subcategory.model_dump())
    return subcategory

@api_router.get("/subcategories", response_model=List[Subcategory])
async def get_subcategories(category_id: Optional[str] = None):
    query = {"category_id": category_id} if category_id else {}
    subcategories = await db.subcategories.find(query, {"_id": 0}).to_list(1000)
    return subcategories

@api_router.put("/subcategories/{subcategory_id}", response_model=Subcategory)
async def update_subcategory(subcategory_id: str, input: SubcategoryCreate):
    result = await db.subcategories.find_one({"id": subcategory_id}, {"_id": 0})
    if not result:
        raise HTTPException(status_code=404, detail="Subcategory not found")
    
    await db.subcategories.update_one(
        {"id": subcategory_id}, 
        {"$set": {"name": input.name, "category_id": input.category_id}}
    )
    result['name'] = input.name
    result['category_id'] = input.category_id
    return Subcategory(**result)

@api_router.delete("/subcategories/{subcategory_id}")
async def delete_subcategory(subcategory_id: str):
    result = await db.subcategories.delete_one({"id": subcategory_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Subcategory not found")
    return {"message": "Subcategory deleted"}


# --- RULES ---
@api_router.post("/rules")
async def create_rule(input: RuleCreate):
    # Upsert by (source, contains, sign) — an inline categorization click for the
    # same concept must not create a duplicate rule; instead update the existing one.
    existing = await db.rules.find_one(
        {"source": input.source, "contains": input.contains, "sign": input.sign},
        {"_id": 0},
    )

    if existing:
        await db.rules.update_one(
            {"id": existing["id"]},
            {"$set": {
                "category_id": input.category_id,
                "subcategory_id": input.subcategory_id,
                "priority": input.priority,
                "active": input.active,
            }},
        )
        rule_doc = {**existing,
                    "category_id": input.category_id,
                    "subcategory_id": input.subcategory_id,
                    "priority": input.priority,
                    "active": input.active}
        created = False
    else:
        rule = Rule(**input.model_dump())
        await db.rules.insert_one(rule.model_dump())
        rule_doc = rule.model_dump()
        created = True

    # Apply rule to existing transactions (same logic for both create and update paths)
    updated_count = 0

    if input.source == 'bank':
        # Find matching transactions
        query = {}
        
        # Match concept (case insensitive)
        if input.contains:
            query['concept'] = {'$regex': re.escape(input.contains), '$options': 'i'}
        
        # Match sign if specified
        if input.sign == '+':
            query['amount'] = {'$gt': 0}
        elif input.sign == '-':
            query['amount'] = {'$lt': 0}
        
        # Update all matching transactions
        result = await db.transactions.update_many(
            query,
            {'$set': {
                'category_id': input.category_id,
                'subcategory_id': input.subcategory_id
            }}
        )
        updated_count = result.modified_count
        
    elif input.source == 'card':
        # Find matching card transactions
        query = {}
        
        if input.contains:
            query['concept'] = {'$regex': re.escape(input.contains), '$options': 'i'}
        
        if input.sign == '+':
            query['amount'] = {'$gt': 0}
        elif input.sign == '-':
            query['amount'] = {'$lt': 0}
        
        # Update all matching card transactions
        result = await db.card_transactions.update_many(
            query,
            {'$set': {
                'category_id': input.category_id,
                'subcategory_id': input.subcategory_id
            }}
        )
        updated_count = result.modified_count

    action = 'creada' if created else 'actualizada'
    return {
        'rule': rule_doc,
        'applied_to_existing': updated_count,
        'created': created,
        'message': f'Regla {action} y aplicada a {updated_count} movimiento(s) existente(s)'
    }

@api_router.get("/rules", response_model=List[Rule])
async def get_rules():
    rules = await db.rules.find({}, {"_id": 0}).sort('priority', -1).to_list(1000)
    return rules

@api_router.delete("/rules/{rule_id}")
async def delete_rule(rule_id: str):
    result = await db.rules.delete_one({"id": rule_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Rule not found")
    return {"message": "Rule deleted"}

@api_router.put("/rules/{rule_id}")
async def update_rule(rule_id: str, input: RuleCreate):
    result = await db.rules.find_one({"id": rule_id}, {"_id": 0})
    if not result:
        raise HTTPException(status_code=404, detail="Rule not found")
    
    updated_data = input.model_dump()
    updated_data['id'] = rule_id
    await db.rules.replace_one({"id": rule_id}, updated_data)
    
    # Apply updated rule to existing transactions
    updated_count = 0
    
    if input.source == 'bank':
        query = {}
        if input.contains:
            query['concept'] = {'$regex': re.escape(input.contains), '$options': 'i'}
        if input.sign == '+':
            query['amount'] = {'$gt': 0}
        elif input.sign == '-':
            query['amount'] = {'$lt': 0}
        
        result = await db.transactions.update_many(
            query,
            {'$set': {
                'category_id': input.category_id,
                'subcategory_id': input.subcategory_id
            }}
        )
        updated_count = result.modified_count
        
    elif input.source == 'card':
        query = {}
        if input.contains:
            query['concept'] = {'$regex': re.escape(input.contains), '$options': 'i'}
        if input.sign == '+':
            query['amount'] = {'$gt': 0}
        elif input.sign == '-':
            query['amount'] = {'$lt': 0}
        
        result = await db.card_transactions.update_many(
            query,
            {'$set': {
                'category_id': input.category_id,
                'subcategory_id': input.subcategory_id
            }}
        )
        updated_count = result.modified_count
    
    return {
        'rule': updated_data,
        'applied_to_existing': updated_count,
        'message': f'Regla actualizada y aplicada a {updated_count} movimiento(s) existente(s)'
    }


# --- TRANSACTIONS ---
@api_router.post("/transactions", response_model=Transaction)
async def create_transaction(input: TransactionCreate):
    dedup_hash = generate_dedup_hash(input.account_id, input.date, input.concept, input.amount)
    
    # Check for duplicates
    existing = await db.transactions.find_one({"dedup_hash": dedup_hash})
    if existing:
        raise HTTPException(status_code=400, detail="Duplicate transaction")
    
    transaction = Transaction(**input.model_dump(), dedup_hash=dedup_hash)
    doc = transaction.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.transactions.insert_one(doc)
    return transaction

@api_router.get("/transactions", response_model=List[Transaction])
async def get_transactions(
    account_id: Optional[str] = None,
    type: Optional[str] = None,
    category_id: Optional[str] = None,
    subcategory_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None
):
    query = {}
    if account_id:
        query['account_id'] = account_id
    if type:
        query['type'] = type
    if category_id:
        query['category_id'] = category_id
    if subcategory_id:
        query['subcategory_id'] = subcategory_id
    if date_from:
        query.setdefault('date', {})['$gte'] = date_from
    if date_to:
        query.setdefault('date', {})['$lte'] = date_to
    
    transactions = await db.transactions.find(query, {"_id": 0}).sort('date', -1).to_list(10000)
    for tx in transactions:
        if isinstance(tx.get('created_at'), str):
            tx['created_at'] = datetime.fromisoformat(tx['created_at'])
    return transactions

@api_router.put("/transactions/{transaction_id}", response_model=Transaction)
async def update_transaction(
    transaction_id: str,
    category_id: Optional[str] = None,
    subcategory_id: Optional[str] = None,
    clear_category: bool = False,
    clear_subcategory: bool = False,
):
    result = await db.transactions.find_one({"id": transaction_id}, {"_id": 0})
    if not result:
        raise HTTPException(status_code=404, detail="Transaction not found")

    update_data = {}
    if clear_category:
        # Clearing the category cascades to subcategory
        update_data['category_id'] = None
        update_data['subcategory_id'] = None
    else:
        if category_id is not None:
            update_data['category_id'] = category_id
        if clear_subcategory:
            update_data['subcategory_id'] = None
        elif subcategory_id is not None:
            update_data['subcategory_id'] = subcategory_id

    if update_data:
        await db.transactions.update_one({"id": transaction_id}, {"$set": update_data})
        result.update(update_data)
    if isinstance(result.get('created_at'), str):
        result['created_at'] = datetime.fromisoformat(result['created_at'])
    return Transaction(**result)


# --- CARD TRANSACTIONS ---
@api_router.post("/card-transactions", response_model=CardTransaction)
async def create_card_transaction(input: CardTransactionCreate):
    dedup_hash = generate_dedup_hash(input.card_id, input.date, input.concept, input.amount)
    
    # Check for duplicates
    existing = await db.card_transactions.find_one({"dedup_hash": dedup_hash})
    if existing:
        raise HTTPException(status_code=400, detail="Duplicate transaction")
    
    transaction = CardTransaction(**input.model_dump(), dedup_hash=dedup_hash)
    doc = transaction.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.card_transactions.insert_one(doc)
    return transaction

@api_router.get("/card-transactions", response_model=List[CardTransaction])
async def get_card_transactions(
    card_id: Optional[str] = None,
    category_id: Optional[str] = None,
    subcategory_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None
):
    query = {}
    if card_id:
        query['card_id'] = card_id
    if category_id:
        query['category_id'] = category_id
    if subcategory_id:
        query['subcategory_id'] = subcategory_id
    if date_from:
        query.setdefault('date', {})['$gte'] = date_from
    if date_to:
        query.setdefault('date', {})['$lte'] = date_to
    
    transactions = await db.card_transactions.find(query, {"_id": 0}).sort('date', -1).to_list(10000)
    for tx in transactions:
        if isinstance(tx.get('created_at'), str):
            tx['created_at'] = datetime.fromisoformat(tx['created_at'])
    return transactions

@api_router.put("/card-transactions/{transaction_id}", response_model=CardTransaction)
async def update_card_transaction(
    transaction_id: str,
    category_id: Optional[str] = None,
    subcategory_id: Optional[str] = None,
    clear_category: bool = False,
    clear_subcategory: bool = False,
):
    result = await db.card_transactions.find_one({"id": transaction_id}, {"_id": 0})
    if not result:
        raise HTTPException(status_code=404, detail="Transaction not found")

    update_data = {}
    if clear_category:
        update_data['category_id'] = None
        update_data['subcategory_id'] = None
    else:
        if category_id is not None:
            update_data['category_id'] = category_id
        if clear_subcategory:
            update_data['subcategory_id'] = None
        elif subcategory_id is not None:
            update_data['subcategory_id'] = subcategory_id

    if update_data:
        await db.card_transactions.update_one({"id": transaction_id}, {"$set": update_data})
        result.update(update_data)
    if isinstance(result.get('created_at'), str):
        result['created_at'] = datetime.fromisoformat(result['created_at'])
    return CardTransaction(**result)


# --- IMPORT CSV/XLSX HELPERS ---

def _parse_amount(amount_str: str) -> float:
    """Parse amount handling both CSV (1.234,56) and Excel (1234.56) formats"""
    try:
        if ',' in amount_str and '.' in amount_str:
            amount_str = amount_str.replace('.', '').replace(',', '.')
        elif ',' in amount_str:
            amount_str = amount_str.replace(',', '.')
        return float(amount_str)
    except (ValueError, TypeError):
        return 0.0


def _parse_xlsx(content: bytes) -> tuple:
    """Parse XLSX (Excel 2007+). Returns (rows, headers)."""
    rows = []
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    headers = [cell.value.strip().lower() if cell.value else '' for cell in ws[1]]

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        row_list = list(row) + [None] * (3 - len(row))
        if not row_list[0] and not row_list[1] and not row_list[2]:
            continue

        fecha_val = row_list[0]
        if isinstance(fecha_val, datetime):
            fecha_str = fecha_val.strftime('%d/%m/%Y')
        else:
            fecha_str = str(fecha_val).strip() if fecha_val is not None else ''

        rows.append({
            'fecha': fecha_str,
            'concepto': str(row_list[1]).strip() if row_list[1] is not None else '',
            'importe': str(row_list[2]) if row_list[2] is not None else '0',
        })

    wb.close()
    return rows, headers


def _xls_date_string(cell, datemode) -> str:
    """Convert an xlrd date cell to DD/MM/YYYY, otherwise stringify."""
    if cell.ctype == xlrd.XL_CELL_DATE:
        d = xlrd.xldate_as_tuple(cell.value, datemode)
        return f"{d[2]:02d}/{d[1]:02d}/{d[0]}"
    return str(cell.value).strip() if cell.value else ''


def _parse_xls(content: bytes) -> tuple:
    """Parse legacy XLS (Excel 97-2003). Returns (rows, headers)."""
    rows = []
    wb = xlrd.open_workbook(file_contents=content)
    ws = wb.sheet_by_index(0)

    headers = [str(cell.value).strip().lower() if cell.value else '' for cell in ws.row(0)]

    for row_idx in range(1, ws.nrows):
        if ws.ncols < 2:
            continue
        fecha_val = _xls_date_string(ws.cell(row_idx, 0), wb.datemode) if ws.ncols > 0 else ''
        concepto_val = ''
        if ws.ncols > 1:
            concepto_cell = ws.cell(row_idx, 1)
            concepto_val = str(concepto_cell.value).strip() if concepto_cell.value else ''
        importe_val = '0'
        if ws.ncols > 2:
            importe_cell = ws.cell(row_idx, 2)
            importe_val = str(importe_cell.value) if importe_cell.value else '0'

        if not fecha_val and not concepto_val and importe_val == '0':
            continue

        rows.append({'fecha': fecha_val, 'concepto': concepto_val, 'importe': importe_val})

    return rows, headers


def _parse_csv_text(content: bytes) -> tuple:
    """Parse CSV bytes. Returns (rows, headers)."""
    text_content = content.decode('utf-8')
    separator = detect_separator(text_content)
    reader = csv.DictReader(io.StringIO(text_content), delimiter=separator)
    rows = [{k.strip().lower(): v.strip() for k, v in row.items()} for row in reader]
    return rows, []


async def _build_preview_row(row: dict, source: str, include_type: bool) -> dict:
    """Turn a parsed source row into a preview row with rule categorization."""
    date = normalize_date(row.get('fecha', ''))
    concept = row.get('concepto', '')
    amount = _parse_amount(row.get('importe', '0'))
    category_id, subcategory_id = await apply_rules(concept, amount, source)

    preview = {
        'date': date,
        'concept': concept,
        'amount': amount,
        'category_id': category_id,
        'subcategory_id': subcategory_id,
    }
    if include_type:
        preview['type'] = 'income' if amount > 0 else 'expense'
    return preview


# --- IMPORT CSV/XLSX ---
@api_router.post("/import/parse-csv")
async def parse_csv(file: UploadFile = File(...), import_type: str = Query(...), entity_id: str = Query(...)):
    """Parse CSV or XLSX and return preview with auto-categorization"""
    content = await file.read()
    filename = file.filename.lower()

    try:
        if filename.endswith('.xlsx'):
            rows, detected_columns = _parse_xlsx(content)
        elif filename.endswith('.xls'):
            rows, detected_columns = _parse_xls(content)
        elif filename.endswith('.csv'):
            rows, detected_columns = _parse_csv_text(content)
        else:
            raise HTTPException(status_code=400, detail="Formato de archivo no soportado. Use .csv, .xls o .xlsx")
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error parsing file {filename}: {e}")
        raise HTTPException(status_code=400, detail=f"Error al procesar archivo: {str(e)}")

    logging.info(f"Parsed {len(rows)} rows from file: {filename}")
    if rows:
        logging.info(f"First row sample: {rows[0]}")

    source = 'bank' if import_type == 'account' else 'card'
    include_type = import_type == 'account'
    preview = [await _build_preview_row(row, source, include_type) for row in rows]

    return {
        'preview': preview,
        'count': len(preview),
        'entity_id': entity_id,
        'import_type': import_type,
        'detected_columns': detected_columns,
    }

@api_router.post("/import/confirm")
async def confirm_import(data: Dict[str, Any]):
    """Confirm and save transactions from preview"""
    import_type = data.get('import_type')
    entity_id = data.get('entity_id')
    transactions = data.get('transactions', [])
    
    inserted = 0
    duplicates = 0
    errors = 0
    
    for tx in transactions:
        try:
            if import_type == 'account':
                dedup_hash = generate_dedup_hash(entity_id, tx['date'], tx['concept'], tx['amount'])
                existing = await db.transactions.find_one({"dedup_hash": dedup_hash})
                if existing:
                    duplicates += 1
                    continue
                
                transaction = Transaction(
                    account_id=entity_id,
                    date=tx['date'],
                    concept=tx['concept'],
                    amount=tx['amount'],
                    type=tx['type'],
                    category_id=tx.get('category_id'),
                    subcategory_id=tx.get('subcategory_id'),
                    dedup_hash=dedup_hash
                )
                doc = transaction.model_dump()
                doc['created_at'] = doc['created_at'].isoformat()
                await db.transactions.insert_one(doc)
                inserted += 1
            
            elif import_type == 'card':
                dedup_hash = generate_dedup_hash(entity_id, tx['date'], tx['concept'], tx['amount'])
                existing = await db.card_transactions.find_one({"dedup_hash": dedup_hash})
                if existing:
                    duplicates += 1
                    continue
                
                transaction = CardTransaction(
                    card_id=entity_id,
                    date=tx['date'],
                    concept=tx['concept'],
                    amount=tx['amount'],
                    category_id=tx.get('category_id'),
                    subcategory_id=tx.get('subcategory_id'),
                    dedup_hash=dedup_hash
                )
                doc = transaction.model_dump()
                doc['created_at'] = doc['created_at'].isoformat()
                await db.card_transactions.insert_one(doc)
                inserted += 1
        except Exception as e:
            errors += 1
            logging.error(f"Error importing transaction: {e}")
    
    return {
        'inserted': inserted,
        'duplicates': duplicates,
        'errors': errors
    }


# --- DASHBOARD ---
@api_router.get("/dashboard")
async def get_dashboard(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    account_id: Optional[str] = None,
    type: Optional[str] = None,
    category_id: Optional[str] = None
):
    """Get dashboard KPIs"""
    query = {}
    if account_id:
        query['account_id'] = account_id
    if type:
        query['type'] = type
    if category_id:
        query['category_id'] = category_id
    if date_from:
        query.setdefault('date', {})['$gte'] = date_from
    if date_to:
        query.setdefault('date', {})['$lte'] = date_to
    
    transactions = await db.transactions.find(query, {"_id": 0}).to_list(10000)
    
    total_income = sum(tx['amount'] for tx in transactions if tx['type'] == 'income')
    total_expense = sum(abs(tx['amount']) for tx in transactions if tx['type'] == 'expense')
    net_flow = total_income - total_expense
    
    return {
        'total_income': total_income,
        'total_expense': total_expense,
        'net_flow': net_flow,
        'transaction_count': len(transactions)
    }


@api_router.get("/dashboard/expense-by-category")
async def get_expense_by_category(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    account_id: Optional[str] = None,
):
    """Aggregate expenses per category (bank + card) for donut chart."""
    query = {'type': 'expense'}
    if account_id:
        query['account_id'] = account_id
    if date_from:
        query.setdefault('date', {})['$gte'] = date_from
    if date_to:
        query.setdefault('date', {})['$lte'] = date_to

    txs = await db.transactions.find(query, {"_id": 0}).to_list(50000)

    card_query = {}
    if date_from:
        card_query.setdefault('date', {})['$gte'] = date_from
    if date_to:
        card_query.setdefault('date', {})['$lte'] = date_to
    # Card expenses are the negative-amount card_transactions
    card_query['amount'] = {'$lt': 0}
    card_txs = await db.card_transactions.find(card_query, {"_id": 0}).to_list(50000)

    totals: Dict[str, float] = {}
    for tx in txs + card_txs:
        key = tx.get('category_id') or '__uncategorized__'
        totals[key] = totals.get(key, 0.0) + abs(tx['amount'])

    categories = await db.categories.find({}, {"_id": 0}).to_list(1000)
    name_by_id = {c['id']: c['name'] for c in categories}

    result = [
        {
            'category_id': None if cat_id == '__uncategorized__' else cat_id,
            'name': 'Sin categoría' if cat_id == '__uncategorized__' else name_by_id.get(cat_id, 'Sin categoría'),
            'total': round(total, 2),
        }
        for cat_id, total in totals.items()
    ]
    result.sort(key=lambda r: r['total'], reverse=True)
    return result


@api_router.get("/dashboard/monthly-summary")
async def get_monthly_summary(
    year: int,
    account_id: Optional[str] = None,
):
    """Return 12 monthly buckets with income, expense and net_flow for the given year."""
    query = {}
    if account_id:
        query['account_id'] = account_id
    query['date'] = {'$gte': f'{year}-01-01', '$lte': f'{year}-12-31'}

    txs = await db.transactions.find(query, {"_id": 0}).to_list(100000)

    buckets = [{'month': m, 'income': 0.0, 'expense': 0.0, 'net_flow': 0.0} for m in range(1, 13)]
    for tx in txs:
        try:
            month = int(tx['date'].split('-')[1])
        except (ValueError, IndexError, KeyError):
            continue
        if not 1 <= month <= 12:
            continue
        b = buckets[month - 1]
        if tx.get('type') == 'income':
            b['income'] += tx['amount']
        elif tx.get('type') == 'expense':
            b['expense'] += abs(tx['amount'])

    month_labels = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
    for b in buckets:
        b['income'] = round(b['income'], 2)
        b['expense'] = round(b['expense'], 2)
        b['net_flow'] = round(b['income'] - b['expense'], 2)
        b['label'] = month_labels[b['month'] - 1]

    return buckets


# --- SETTINGS ---
@api_router.get("/settings", response_model=Settings)
async def get_settings():
    settings = await db.settings.find_one({}, {"_id": 0})
    if not settings:
        # Create default settings
        default = Settings()
        await db.settings.insert_one(default.model_dump())
        return default
    return Settings(**settings)

@api_router.put("/settings", response_model=Settings)
async def update_settings(input: SettingsUpdate):
    settings = await db.settings.find_one({}, {"_id": 0})
    if not settings:
        new_settings = Settings(**input.model_dump())
        await db.settings.insert_one(new_settings.model_dump())
        return new_settings
    
    await db.settings.update_one({}, {"$set": input.model_dump()})
    settings.update(input.model_dump())
    return Settings(**settings)


# Include the router in the main app
app.include_router(auth_router)
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
